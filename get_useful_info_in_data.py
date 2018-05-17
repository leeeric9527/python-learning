#-*- encoding:UTF-8 -*-

import numpy as np 
import pandas as pd 
import openpyxl as op
import regex as re
#载入分析数据
book=op.load_workbook('data.xlsx','utf-8')
isheet=book.worksheets[0]
rows=isheet.max_row
cols=isheet.max_column
#输出分析数据的表格
wb=op.Workbook()
osheet=wb.worksheets[0]
#存储一个 属性名 --> 属性列下标 的映射
attribute_name={'domain_name':1,'legal_type':2,'Referer':3,'POST':4,'Cookie':5,'Content-Length':6}
osheet.cell(row=1,column=1).value='domain_name'
osheet.cell(row=1,column=2).value='legal_type'
osheet.cell(row=1,column=3).value='Referer'
osheet.cell(row=1,column=4).value='POST'
osheet.cell(row=1,column=5).value='Cookie'
osheet.cell(row=1,column=6).value='Content-Length'

for i in range(2,rows+1):#向字典里添加元素  rows+1
	#将域名和类型写入表格
	osheet.cell(row=i,column=1).value=isheet.cell(row=i,column=1).value
	osheet.cell(row=i,column=2).value=isheet.cell(row=i,column=3).value
	pst=isheet.cell(row=i,column=2).value
	pst=pst.replace('_x000D_','')#去掉无用字符
	#使用正则匹配从post报文中提取出有用信息
	ref_pattern='Re[fb]erer[^\n]+\n'
	ref=re.search(ref_pattern,pst)
	if ref!=None:
		ref=ref.group()
		ref=ref[8:].strip()[7:]
		# index=ref.find('/')
		# if index>=0:
		# 	ref=ref[index:]
		# ref=ref.replace('/',' ').replace('&',' ').replace('=',' ').replace('.',' ')
		# index=ref.find('?')
		# method=''
		# if index>=0:# path + method 模式
		# 	[path, method]=[ref[0:index], ref[index+1:]]
		# 	path=path.strip()
		# 	method=method.split(' ')[::2]#获取所有方法名称
		# else:# path 模式
		# 	path=ref.strip()
		osheet.cell(row=i,column=3).value=ref#path+" ".join(method)

	pas_pattern='POST[^\n]+\n'
	pas=re.search(pas_pattern,pst)
	if pas!=None:
		pas=pas.group()
		pas=pas[4:-9].strip()
		index=pas.find('/')
		if index>=0:
			pas=pas[index:]
		osheet.cell(row=i,column=4).value=pas

	cok_pattern='[Cc]o[ok]kie:([^=@\n]+[=][^=@\n])+'
	cok=re.search(cok_pattern,pst)
	if cok!=None:
		cok=cok.group()
		#cok=cok[7:].strip()
		osheet.cell(row=i,column=5).value=cok
	
	col_pattern='[Cc]on[tp]ent-[Ll]en[gc]th[^\n]+[\n]{1}'
	col=re.search(col_pattern,pst)
	if col!=None:
		col=col.group()
		col=col[15:].strip('[\n ]')
		osheet.cell(row=i,column=6).value=col
wb.save('Relative-data.xlsx')