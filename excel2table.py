#-*- encoding: UTF-8 -*-

import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import numpy as np
import pandas as pd
import regex as re
#载入分析数据
book=load_workbook('exdata.xlsx','utf-8')
sheet=book.worksheets[0]
rows=sheet.max_row
cols=sheet.max_column
#输出分析数据的表格
wb=Workbook()
osheet=wb.worksheets[0]
attribute_name={'域名': 1, 'type': 2}#存储一个 属性名 --> 属性列下标 的映射
attribute_num=2
osheet.cell(row=1,column=1).value='domain_name'
osheet.cell(row=1,column=2).value='legal_type'
for i in range(2,rows+1):#向字典里添加元素  rows+1
	#将域名和类型写入表格
	osheet.cell(row=i,column=1).value=sheet.cell(row=i,column=1).value
	osheet.cell(row=i,column=2).value=sheet.cell(row=i,column=3).value
	#对POST报文进行处理
	pval=sheet.cell(i,2).value
	#去除无用字符
	pval=pval.replace('_x000D_','')
	#筛除无用密文------也可能不是
	pval=pval.split('\n\n')
	att_val=pval[0]
	#生成<属性，值>对
	att_val=att_val.split('\n')
	#print(att_val)#测试使用
	for j in range(0,len(att_val)):
		#分割字符串为 属性，值 对
		if j==0:
			index=att_val[j].find(' ')
			[attr,value]=[att_val[j][0:index].strip(),att_val[j][index+1:].strip()]
		else:
			index=att_val[j].find(':')
			[attr,value]=[att_val[j][0:index].strip(),att_val[j][index+1:].strip()]
		#筛除值为空的项目
		if len(value)==0:
			continue
		#属性列表里没有该属性名时增加属性
		if attribute_name.get(attr.upper())==None:
			attribute_name[attr.upper()]=attribute_num+1
			attribute_num+=1
			osheet.cell(row=1,column=attribute_name[attr.upper()]).value=attr
		#写入输出的表格，直接写单元格就行
		osheet.cell(row=i,column=attribute_name[attr.upper()]).value=value
wb.save('table.xlsx')
	