#-*- encoding:UTF-8 -*-
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import numpy as np
import pandas as pd
import re
strinfo = re.compile('[ ]+')
book=load_workbook('ex2.xlsx','utf-8')
sheet=book.worksheets[0]
rows=sheet.max_row
cols=sheet.max_column
Post={}
Type={}
for i in range(2,rows+1):#向字典里添加元素
	dn=sheet.cell(i,1).value
	pv=sheet.cell(i,2).value
	tv=sheet.cell(i,3).value
	if Post.get(dn)==None:#第一次遇到這个域名
		Post[dn]=pv
		Type[dn]=tv
	else:
		Post[dn]+='\n'+pv

wb=Workbook()
sh=wb.worksheets[0]#输出表格
for i in range(2,rows+1):#从字典中取出内容存入excel
	dn=sheet.cell(i,1).value
	if i==2:
		Post[dn]=Post[dn].replace('/',' ').replace(':',' ')
		Post[dn]=Post[dn].replace('(',' ').replace(')',' ')
		Post[dn]=Post[dn].replace('*',' ').replace(';',' ')
		Post[dn]=Post[dn].replace('\t',' ').replace('\n',' ')
		Post[dn]=Post[dn].replace('$',' ').replace('@',' ')
		Post[dn]=Post[dn].replace('=',' ').replace('&',' ')
		Post[dn]=Post[dn].replace(',',' ').replace('?',' ')
		Post[dn]=strinfo.sub(' ',Post[dn])
		sh.append([dn,Post[dn],Type[dn]])
	else:
		if dn!=sheet.cell(i-1,1).value:
			Post[dn]=Post[dn].replace('/',' ').replace(':',' ')
			Post[dn]=Post[dn].replace('(',' ').replace(')',' ')
			Post[dn]=Post[dn].replace('*',' ').replace(';',' ')
			Post[dn]=Post[dn].replace('\t',' ').replace('\n',' ')
			Post[dn]=Post[dn].replace('$',' ').replace('@',' ')
			Post[dn]=Post[dn].replace('=',' ').replace('&',' ')
			Post[dn]=Post[dn].replace(',',' ').replace('?',' ')
			Post[dn]=strinfo.sub(' ',Post[dn])
			sh.append([dn,Post[dn],Type[dn]])
		else:
			continue
replace('_x000D_','')
wb.save('out.csv')