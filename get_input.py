#-*- encoding:UTF-8 -*-

import numpy as np 
import pandas as pd 
import openpyxl as op
import regex as re
#载入分析数据
book=op.load_workbook('feature_data.xlsx','utf-8')
isheet=book.worksheets[0]
rows=isheet.max_row
cols=isheet.max_column#3 4 5
#输出分析数据的表格
wb=op.Workbook()
osheet=wb.worksheets[0]
osheet.cell(row=1,column=1).value='domain_name'
osheet.cell(row=1,column=2).value='legal_type'
osheet.cell(row=1,column=3).value='Path'
osheet.cell(row=1,column=4).value='Method'
osheet.cell(row=1,column=5).value='Cookie'
osheet.cell(row=1,column=6).value='Content-Length'

path={}
method={}
cookie={}
dindex={'domain_name':1}
domain_num=1
for i in range(2,rows+1):#向字典里添加元素  rows+1
	#将域名和类型写入表格
	if dindex.get(isheet.cell(i,1).value)==None:#为第一个域名或者新域名
		if len(path)>1:#为进入下一个域名
			osheet.cell(row=domain_num,column=3).value=str(path).replace('None','')
			osheet.cell(row=domain_num,column=4).value=str(method).replace('None','')
			osheet.cell(row=domain_num,column=5).value=str(cookie).replace('None','')
			osheet.cell(row=domain_num,column=1).value=isheet.cell(i-1,1).value
			osheet.cell(row=domain_num,column=2).value=isheet.cell(i-1,2).value
			osheet.cell(row=domain_num,column=6).value=isheet.cell(i-1,6).value
			path={}
			method={}
			cookie={}

		#新域名
		domain_num+=1
		dindex[isheet.cell(i,1).value]=domain_num

		path=set(str(isheet.cell(i,3).value).split())
		method=set(str(isheet.cell(i,4).value).split())
		cookie=set(str(isheet.cell(i,5).value).split())	
	else:
		path.update(str(isheet.cell(i,3).value).split())
		method.update(str(isheet.cell(i,4).value).split())
		cookie.update(str(isheet.cell(i,5).value).split())
wb.save('input.xlsx')